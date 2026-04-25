"""Admin chatbot service — AI-powered student filtering & report generation."""

import json
import re
from collections import defaultdict
from io import BytesIO

from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models.collections import (
    GROUP_TESTS,
    GROUP_TEST_RESULTS,
    JOB_DESCRIPTIONS,
    SKILLS,
    TOPICS,
    USERS,
)
from services.group_test_service import _refresh_topic_statuses
from utils.gemini import call_gemini
from utils.helpers import str_objectids


# ─── Gemini query parser ─────────────────────────────────────────────────────

async def _parse_query(query: str, group_tests: list[dict], jd_content: str | None) -> dict:
    """Ask Gemini to extract structured filter parameters from a natural-language query."""
    gt_list = [{"id": gt["id"], "name": gt["name"]} for gt in group_tests]

    jd_context = ""
    if jd_content:
        jd_context = (
            f"\n\nJob Description:\n{jd_content}\n"
            "Use this JD to rank students by skill relevance when use_jd_ranking is true."
        )

    prompt = (
        f'Admin query: "{query}"\n\n'
        f"Available group tests: {json.dumps(gt_list)}"
        f"{jd_context}\n\n"
        "Extract filter parameters and return ONLY a JSON object (no markdown, no extra text):\n"
        "{\n"
        '  "group_test_id": "<id from the list, or null if none matches>",\n'
        '  "group_test_name": "<matched name or null>",\n'
        '  "top_k": <integer or null>,\n'
        '  "min_score": <number 0-100 or null>,\n'
        '  "use_jd_ranking": <true if JD was provided and should influence ranking>,\n'
        '  "response_message": "<short 1-2 sentence message describing the filter result>"\n'
        "}\n\n"
        "Rules:\n"
        "- Match group_test_id to the best-fitting group test. null = show all students across all tests.\n"
        "- top_k: number from phrases like 'top 5', 'top k', 'best 10'. null = all.\n"
        "- min_score: extract from 'score above 70', 'minimum 80%'. null = no filter.\n"
        "- response_message: friendly description of what was filtered.\n"
        "Return ONLY valid JSON, no other text."
    )

    raw = await call_gemini(prompt)
    cleaned = raw.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"```[a-z]*\n?", "", cleaned).strip().rstrip("`").strip()
    try:
        return json.loads(cleaned)
    except Exception:
        # Fallback: return empty params so the caller can show a helpful message
        return {
            "group_test_id": None,
            "group_test_name": None,
            "top_k": None,
            "min_score": None,
            "use_jd_ranking": False,
            "response_message": "I couldn't understand the query. Please try something like: 'top 5 students in SWE group'.",
        }


# ─── Data helpers ─────────────────────────────────────────────────────────────

async def _user_info(user_id: str, db) -> dict:
    try:
        user = await db[USERS].find_one({"_id": ObjectId(user_id)})
    except Exception:
        user = None
    if not user:
        return {"reg_no": "N/A", "name": "", "email": ""}
    return {
        "reg_no": user.get("reg_no") or "N/A",
        "name": user.get("name", ""),
        "email": user.get("email", ""),
    }


async def _jd_skills(jd_id: str, db) -> tuple[str | None, list[str]]:
    """Return (jd_content_str, required_skills_list)."""
    try:
        doc = await db[JOB_DESCRIPTIONS].find_one({"_id": ObjectId(jd_id)})
    except Exception:
        doc = None
    if not doc:
        return None, []
    content = f"Title: {doc.get('title', '')}\n{doc.get('description', '')}"
    return content, doc.get("required_skills") or []


def _skill_match_pct(student_skills: list[str], jd_skills: list[str]) -> float | None:
    if not jd_skills:
        return None
    s_lower = [s.lower() for s in student_skills]
    j_lower = [s.lower() for s in jd_skills]
    matched = sum(1 for j in j_lower if any(j in s or s in j for s in s_lower))
    return round(matched / len(j_lower) * 100, 1)


# ─── Main query processor ────────────────────────────────────────────────────

async def process_chatbot_query(query: str, jd_id: str | None) -> dict:
    """Parse admin query, aggregate student data, apply filters, return ranked rows."""
    db = get_db()

    # Fetch all group tests
    gt_cursor = db[GROUP_TESTS].find({}).sort("created_at", -1)
    gt_docs = await gt_cursor.to_list(length=300)
    all_group_tests = [
        {
            "id": str(d["_id"]),
            "name": d.get("name", ""),
            "topic_ids": d.get("topic_ids") or [],
        }
        for d in gt_docs
    ]

    # Fetch JD if provided
    jd_content, jd_req_skills = (None, [])
    if jd_id:
        jd_content, jd_req_skills = await _jd_skills(jd_id, db)

    # Let Gemini parse the query
    parsed = await _parse_query(query, all_group_tests, jd_content)

    group_test_id: str | None = parsed.get("group_test_id")
    top_k: int | None = parsed.get("top_k")
    min_score: float | None = parsed.get("min_score")
    use_jd_ranking: bool = bool(parsed.get("use_jd_ranking")) and bool(jd_req_skills)
    response_message: str = parsed.get("response_message") or "Here are the filtered results."

    # Build topic column list from the matched group test
    topic_columns: list[dict] = []
    group_test_name: str = ""

    if group_test_id:
        gt_doc = next((g for g in all_group_tests if g["id"] == group_test_id), None)
        if gt_doc:
            group_test_name = gt_doc["name"]
            for tid in gt_doc["topic_ids"]:
                try:
                    t = await db[TOPICS].find_one({"_id": ObjectId(tid)})
                except Exception:
                    t = None
                if t:
                    topic_columns.append({"id": tid, "name": t.get("name", tid)})

    # Fetch relevant results
    results_filter = {"group_test_id": group_test_id} if group_test_id else {}
    results_cursor = db[GROUP_TEST_RESULTS].find(results_filter)
    results_docs = await results_cursor.to_list(length=2000)

    # Group by user_id; pick best attempt per user per group_test
    # Key: (user_id, group_test_id) → list of attempts
    attempts_map: dict[tuple, list] = defaultdict(list)
    for r in results_docs:
        key = (r.get("user_id", ""), r.get("group_test_id", ""))
        attempts_map[key].append(r)

    rows: list[dict] = []
    seen_users: set[str] = set()

    for (uid, gt_id), attempts in attempts_map.items():
        if not uid:
            continue

        # Refresh topic statuses and choose best attempt
        best = None
        for attempt in attempts:
            attempt = await _refresh_topic_statuses(attempt, db)
            score = attempt.get("overall_score") or 0
            if best is None or score > (best.get("overall_score") or 0):
                best = attempt

        user = await _user_info(uid, db)

        # Per-topic scores from best attempt
        topic_scores: dict[str, dict] = {}
        for tr in best.get("topic_results") or []:
            tid = tr.get("topic_id", "")
            topic_scores[tid] = {
                "topic_name": tr.get("topic_name", ""),
                "score": tr.get("overall_score"),
                "status": tr.get("status", "pending"),
            }

        # JD skill match
        skill_match: float | None = None
        if use_jd_ranking:
            skills_doc = await db[SKILLS].find_one({"user_id": uid})
            student_skills = (skills_doc or {}).get("skills") or []
            skill_match = _skill_match_pct(student_skills, jd_req_skills)

        row = {
            "user_id": uid,
            "reg_no": user["reg_no"],
            "name": user["name"],
            "email": user["email"],
            "group_test_id": gt_id,
            "group_test_name": best.get("group_test_name") or group_test_name,
            "overall_score": round(best.get("overall_score") or 0, 1),
            "total_attempts": len(attempts),
            "status": best.get("status", "in_progress"),
            "topic_scores": topic_scores,
            "skill_match": skill_match,
            "rank": 0,  # assigned below
        }
        rows.append(row)

    # If multiple group tests queried (no filter), collect unique topic columns
    if not group_test_id:
        topic_set: dict[str, str] = {}
        for r in rows:
            for tid, ts in r["topic_scores"].items():
                if tid not in topic_set:
                    topic_set[tid] = ts["topic_name"]
        topic_columns = [{"id": tid, "name": name} for tid, name in topic_set.items()]

    # Sort
    if use_jd_ranking:
        rows.sort(
            key=lambda r: (r["skill_match"] or 0) * 0.4 + (r["overall_score"] or 0) * 0.6,
            reverse=True,
        )
    else:
        rows.sort(key=lambda r: r["overall_score"] or 0, reverse=True)

    # Min score filter
    if min_score is not None:
        rows = [r for r in rows if (r["overall_score"] or 0) >= min_score]

    # Assign ranks
    for i, row in enumerate(rows):
        row["rank"] = i + 1

    # Top-k slice
    if top_k and top_k > 0:
        rows = rows[:top_k]

    return {
        "message": response_message,
        "group_test_name": group_test_name or "All Group Tests",
        "group_test_id": group_test_id,
        "topic_columns": topic_columns,
        "rows": rows,
        "total": len(rows),
    }


# ─── Update student ───────────────────────────────────────────────────────────

async def update_student_info(user_id: str, reg_no: str | None, name: str | None) -> dict:
    """Allow admin to correct a student's reg_no or name."""
    db = get_db()
    update: dict = {}
    if reg_no is not None:
        reg_no = reg_no.strip()
        if reg_no:
            # Uniqueness check
            existing = await db[USERS].find_one(
                {"reg_no": reg_no, "_id": {"$ne": ObjectId(user_id)}}
            )
            if existing:
                raise ValueError("This register number is already used by another student.")
            update["reg_no"] = reg_no
    if name is not None:
        name = name.strip()
        if name:
            update["name"] = name
    if not update:
        raise ValueError("Nothing to update.")
    await db[USERS].update_one({"_id": ObjectId(user_id)}, {"$set": update})
    user = await db[USERS].find_one({"_id": ObjectId(user_id)})
    return {
        "user_id": user_id,
        "reg_no": (user or {}).get("reg_no") or "N/A",
        "name": (user or {}).get("name", ""),
        "email": (user or {}).get("email", ""),
    }


# ─── Excel export ─────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_SCORE_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(name="Calibri", size=10)
_BOLD_DATA_FONT = Font(name="Calibri", bold=True, size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _thin_border() -> Border:
    s = Side(style="thin", color="B0BEC5")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_excel(rows: list[dict], topic_columns: list[dict], group_test_name: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    border = _thin_border()

    # ── Header row ────────────────────────────────────────────────────────────
    headers = ["Rank", "Reg No", "Name", "Email"]
    for tc in topic_columns:
        headers.append(f"{tc['name']}\nScore")
    headers += ["Overall\nScore", "Attempts", "Status"]
    if any(r.get("skill_match") is not None for r in rows):
        headers.append("JD Match\n(%)")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = border
    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_num, row in enumerate(rows, 2):
        use_alt = row_num % 2 == 0
        row_fill = _ALT_FILL if use_alt else None

        data: list = [
            row.get("rank", row_num - 1),
            row.get("reg_no", ""),
            row.get("name", ""),
            row.get("email", ""),
        ]

        for tc in topic_columns:
            ts = row.get("topic_scores", {}).get(tc["id"], {})
            score = ts.get("score")
            data.append(f"{score:.1f}%" if score is not None else "—")

        overall = row.get("overall_score")
        data.append(f"{overall:.1f}%" if overall is not None else "—")
        data.append(row.get("total_attempts", 1))
        data.append((row.get("status") or "").replace("_", " ").title())

        if any(r.get("skill_match") is not None for r in rows):
            sm = row.get("skill_match")
            data.append(f"{sm:.1f}%" if sm is not None else "—")

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = border
            cell.font = _DATA_FONT
            if col_idx in (1,):  # rank → bold + centered
                cell.font = _BOLD_DATA_FONT
                cell.alignment = _CENTER
            elif col_idx in (2, 3, 4):
                cell.alignment = _LEFT
            else:
                cell.alignment = _CENTER
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_num].height = 20

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = [6, 16, 22, 28]
    for _ in topic_columns:
        col_widths.append(14)
    col_widths += [14, 10, 14]
    if any(r.get("skill_match") is not None for r in rows):
        col_widths.append(12)

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Title row above headers ────────────────────────────────────────────────
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=f"Student Results — {group_test_name}")
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    title_cell.alignment = _LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 5))
    ws.row_dimensions[1].height = 28

    # ── Freeze header row ──────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
